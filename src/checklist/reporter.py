"""Excel report generator for checklist results."""

from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models import RuleResult


# Style constants
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_NC_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_PASS_FONT = Font(color="006100", bold=True)
_FAIL_FONT = Font(color="9C0006", bold=True)
_NC_FONT = Font(color="1F4E79", bold=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_report(results: list[RuleResult], output_path: str | Path,
                    job_name: str = "", components_top=None, components_bot=None,
                    references_dir: str | Path = None):
    """Generate an Excel checklist report.

    Tabs are ordered: Summary, Details, manage_list, then one tab per rule
    sorted numerically by rule ID (e.g. CKL-01-001, CKL-01-002, CKL-03-010).

    Args:
        results: List of RuleResult objects from the checklist engine
        output_path: Path to write the .xlsx file
        job_name: Job name for the report header
        components_top: List of Component objects on the TOP layer
        components_bot: List of Component objects on the BOTTOM layer
        references_dir: Directory containing managed-parts CSV files
    """
    wb = Workbook()

    # Summary sheet (uses the default first sheet)
    _create_summary_sheet(wb, results, job_name)

    # Detail sheet (all rules in one table)
    _create_detail_sheet(wb, results)

    # Managed parts sheet
    _create_manage_list_sheet(wb, components_top or [], components_bot or [],
                               references_dir)

    # Per-rule detail sheets, sorted by rule ID
    sorted_results = sorted(results, key=_rule_sort_key)
    for result in sorted_results:
        _create_rule_sheet(wb, result)

    # Save
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Checklist report saved: {output_path}")

    # Clean up temporary visualisation images
    _cleanup_images(results)


def _rule_sort_key(result: RuleResult) -> tuple:
    """Extract numeric parts from a rule ID for natural sorting.

    Handles formats like CKL-001, CKL-01-001, CKL-03-010, etc.
    """
    nums = re.findall(r"\d+", result.rule_id)
    return tuple(int(n) for n in nums)


def _create_summary_sheet(wb: Workbook, results: list[RuleResult], job_name: str):
    """Create the summary overview sheet."""
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"ODB++ Design Checklist Report - {job_name}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Stats
    total = len(results)
    passed = sum(1 for r in results if r.passed)
    failed = total - passed

    ws["A3"] = "Total Rules:"
    ws["B3"] = total
    ws["A4"] = "Passed:"
    ws["B4"] = passed
    ws["B4"].font = _PASS_FONT
    ws["A5"] = "Failed:"
    ws["B5"] = failed
    ws["B5"].font = _FAIL_FONT

    # Headers
    headers = ["Rule ID", "Category", "Description", "Status", "Message", "Affected Components"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Data rows (sorted by rule ID for consistent ordering)
    for row_idx, result in enumerate(sorted(results, key=_rule_sort_key), 8):
        ws.cell(row=row_idx, column=1, value=result.rule_id).border = _THIN_BORDER
        ws.cell(row=row_idx, column=2, value=result.category).border = _THIN_BORDER
        ws.cell(row=row_idx, column=3, value=result.description).border = _THIN_BORDER

        status_cell = ws.cell(row=row_idx, column=4, value="PASS" if result.passed else "FAIL")
        status_cell.fill = _PASS_FILL if result.passed else _FAIL_FILL
        status_cell.font = _PASS_FONT if result.passed else _FAIL_FONT
        status_cell.border = _THIN_BORDER
        status_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_idx, column=5, value=result.message).border = _THIN_BORDER
        ws.cell(row=row_idx, column=6,
                value=", ".join(result.affected_components[:20])).border = _THIN_BORDER

    # Auto-fit column widths
    _auto_fit_columns(ws)


def _create_detail_sheet(wb: Workbook, results: list[RuleResult]):
    """Create a detail sheet with expanded rule information."""
    ws = wb.create_sheet("Details")

    headers = ["Rule ID", "Category", "Description", "Status",
               "Message", "Affected Components", "Details"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER

    for row_idx, result in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=result.rule_id).border = _THIN_BORDER
        ws.cell(row=row_idx, column=2, value=result.category).border = _THIN_BORDER
        ws.cell(row=row_idx, column=3, value=result.description).border = _THIN_BORDER

        status_cell = ws.cell(row=row_idx, column=4, value="PASS" if result.passed else "FAIL")
        status_cell.fill = _PASS_FILL if result.passed else _FAIL_FILL
        status_cell.font = _PASS_FONT if result.passed else _FAIL_FONT
        status_cell.border = _THIN_BORDER

        ws.cell(row=row_idx, column=5, value=result.message).border = _THIN_BORDER
        ws.cell(row=row_idx, column=6,
                value="\n".join(result.affected_components)).border = _THIN_BORDER
        ws.cell(row=row_idx, column=6).alignment = Alignment(wrap_text=True)

        # Details as formatted string
        if result.details:
            detail_str = "\n".join(f"{k}: {v}" for k, v in result.details.items())
            ws.cell(row=row_idx, column=7, value=detail_str).border = _THIN_BORDER
            ws.cell(row=row_idx, column=7).alignment = Alignment(wrap_text=True)

    _auto_fit_columns(ws)


def _create_rule_sheet(wb: Workbook, result: RuleResult):
    """Create a dedicated sheet for a single rule's detailed results.

    The tab name is the rule ID (e.g. "CKL-01-001").  The sheet contains
    a header block with rule metadata followed by a results table.

    **Tabular mode** (preferred for new-style checklists):
    If ``result.details`` contains both ``"columns"`` (list of str) and
    ``"rows"`` (list of dicts), the sheet renders a proper multi-column
    table with those headers and colour-coded status cells.

    **Legacy mode**: falls back to the original key-value / list rendering
    for older rule implementations.
    """
    ws = wb.create_sheet(result.rule_id)

    # -- Header block ----------------------------------------------------------
    ws["A1"] = "Rule ID:"
    ws["B1"] = result.rule_id
    ws["A1"].font = Font(bold=True)

    ws["A2"] = "Category:"
    ws["B2"] = result.category
    ws["A2"].font = Font(bold=True)

    ws["A3"] = "Description:"
    ws["B3"] = result.description
    ws["A3"].font = Font(bold=True)

    ws["A4"] = "Status:"
    status_cell = ws["B4"]
    status_cell.value = "PASS" if result.passed else "FAIL"
    status_cell.fill = _PASS_FILL if result.passed else _FAIL_FILL
    status_cell.font = _PASS_FONT if result.passed else _FAIL_FONT

    ws["A5"] = "Message:"
    ws["B5"] = result.message
    ws["A5"].font = Font(bold=True)

    # -- Tabular detail rows ---------------------------------------------------
    columns = result.details.get("columns")
    rows = result.details.get("rows")

    end_row = 6  # after the header block

    if isinstance(columns, list) and isinstance(rows, list):
        _write_tabular_details(ws, columns, rows, start_row=7)
        end_row = 7 + len(rows)  # header row + data rows

        # Optional second table (e.g. signal-layer results for CKL-03-015)
        sig_columns = result.details.get("signal_columns")
        sig_rows = result.details.get("signal_rows")
        if isinstance(sig_columns, list) and isinstance(sig_rows, list) and sig_rows:
            gap_row = end_row + 2  # +1 blank
            _write_tabular_details(ws, sig_columns, sig_rows, start_row=gap_row)
            end_row = gap_row + len(sig_rows)

    else:
        # -- Legacy detail rendering -------------------------------------------
        if result.affected_components:
            ws["A7"] = "Affected Components:"
            ws["A7"].font = Font(bold=True)
            for i, comp_name in enumerate(result.affected_components):
                cell = ws.cell(row=8 + i, column=1, value=comp_name)
                cell.border = _THIN_BORDER

        start_row = 8 + len(result.affected_components) + 2 if result.affected_components else 8

        for key, value in result.details.items():
            if key in ("columns", "rows"):
                continue
            header_cell = ws.cell(row=start_row, column=1, value=key)
            header_cell.font = Font(bold=True, size=11)
            start_row += 1

            if isinstance(value, list):
                hdr_cell = ws.cell(row=start_row, column=1, value="#")
                hdr_cell.fill = _HEADER_FILL
                hdr_cell.font = _HEADER_FONT
                hdr_cell.border = _THIN_BORDER
                hdr_cell = ws.cell(row=start_row, column=2, value="Detail")
                hdr_cell.fill = _HEADER_FILL
                hdr_cell.font = _HEADER_FONT
                hdr_cell.border = _THIN_BORDER
                start_row += 1

                for idx, item in enumerate(value, 1):
                    ws.cell(row=start_row, column=1, value=idx).border = _THIN_BORDER
                    ws.cell(row=start_row, column=2, value=str(item)).border = _THIN_BORDER
                    start_row += 1
            elif isinstance(value, dict):
                hdr_cell = ws.cell(row=start_row, column=1, value="Key")
                hdr_cell.fill = _HEADER_FILL
                hdr_cell.font = _HEADER_FONT
                hdr_cell.border = _THIN_BORDER
                hdr_cell = ws.cell(row=start_row, column=2, value="Value")
                hdr_cell.fill = _HEADER_FILL
                hdr_cell.font = _HEADER_FONT
                hdr_cell.border = _THIN_BORDER
                start_row += 1

                for k, v in value.items():
                    ws.cell(row=start_row, column=1, value=str(k)).border = _THIN_BORDER
                    ws.cell(row=start_row, column=2, value=str(v)).border = _THIN_BORDER
                    start_row += 1
            else:
                ws.cell(row=start_row, column=1, value=str(value)).border = _THIN_BORDER
                start_row += 1

            start_row += 1

        end_row = start_row

    # -- Visualization images --------------------------------------------------
    if result.images:
        _insert_images(ws, result.images, start_row=end_row + 2)

    _auto_fit_columns(ws)


def _insert_images(ws, images: list[dict], start_row: int):
    """Insert visualisation images into the worksheet.

    Each entry in *images* is a dict with keys:
    - ``path``  – Path to a PNG file
    - ``title`` – caption shown above the image
    - ``width`` – desired width in pixels (default 500)
    """
    _IMG_TITLE_FILL = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    row = start_row
    for img_info in images:
        img_path = Path(img_info["path"])
        if not img_path.exists():
            continue

        # Title row
        title = img_info.get("title", img_path.stem)
        title_cell = ws.cell(row=row, column=1, value=title)
        title_cell.font = Font(bold=True, size=11)
        title_cell.fill = _IMG_TITLE_FILL
        row += 1

        # Insert image
        img = XlImage(str(img_path))
        target_width = img_info.get("width", 500)
        scale = target_width / img.width
        img.width = target_width
        img.height = int(img.height * scale)
        ws.add_image(img, f"A{row}")

        # Skip enough rows to accommodate the image (~20 px per row)
        row += max(1, img.height // 20) + 2

    return row


def _cleanup_images(results: list):
    """Remove temporary visualisation image files after the report is saved."""
    dirs_to_remove: set[Path] = set()
    for r in results:
        for img_info in r.images:
            p = Path(img_info["path"])
            if p.exists():
                dirs_to_remove.add(p.parent)
                p.unlink(missing_ok=True)
    for d in dirs_to_remove:
        try:
            if d.exists() and not any(d.iterdir()):
                d.rmdir()
        except OSError:
            pass


def _write_tabular_details(ws, columns: list[str], rows: list[dict],
                           start_row: int = 7):
    """Write a proper multi-column table into *ws*.

    Colour-codes the ``status`` column with PASS/FAIL fills.
    """
    # Column headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_offset, row_data in enumerate(rows, 1):
        current_row = start_row + row_offset
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = _THIN_BORDER

            # Colour-code status column
            if col_name.lower() == "status":
                val_str = str(value).upper()
                if val_str == "PASS":
                    cell.fill = _PASS_FILL
                    cell.font = _PASS_FONT
                elif val_str == "FAIL":
                    cell.fill = _FAIL_FILL
                    cell.font = _FAIL_FONT
                elif val_str == "NC":
                    cell.fill = _NC_FILL
                    cell.font = _NC_FONT
                cell.alignment = Alignment(horizontal="center")


def _load_csv_part_map(csv_path: Path) -> dict[str, str]:
    """Load a CSV and return a mapping of part_name -> size."""
    part_map: dict[str, str] = {}
    try:
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                pn = (row.get("part_name") or "").strip()
                sz = (row.get("size") or "").strip()
                if pn:
                    part_map[pn] = sz
    except Exception:
        pass
    return part_map


def _count_usage(part_names: set[str], components_top: list,
                 components_bot: list) -> dict[str, dict]:
    """For each managed part_name, count how many board components use it.

    Returns a dict: part_name -> {"top": int, "bottom": int, "total": int}
    """
    counts: dict[str, dict] = {pn: {"top": 0, "bottom": 0, "total": 0}
                                for pn in part_names}
    for comp in components_top:
        if comp.part_name in counts:
            counts[comp.part_name]["top"] += 1
            counts[comp.part_name]["total"] += 1
    for comp in components_bot:
        if comp.part_name in counts:
            counts[comp.part_name]["bottom"] += 1
            counts[comp.part_name]["total"] += 1
    return counts


def _create_manage_list_sheet(wb: Workbook, components_top: list,
                               components_bot: list,
                               references_dir):
    """Create the 'manage_list' sheet tracking usage frequency of managed parts.

    Four sections are generated:
    - Capacitors (merged): combined capacitors_10_list + capacitors_41_list
    - Inductors (2S): inductors_2s_list
    - Capacitors 10-type: capacitors_10_list only (dedicated view)
    - Inductors 2S: inductors_2s_list only (dedicated view)

    Each section shows one row per managed part_name with TOP/BOTTOM/Total
    usage counts, sorted by total count descending.
    """
    ws = wb.create_sheet("manage_list")

    ref = Path(references_dir) if references_dir else None
    if not ref or not ref.is_dir():
        ws["A1"] = "No managed-parts CSV files found in references directory."
        return

    # Load the three active reference CSVs
    cap10_map  = _load_csv_part_map(ref / "capacitors_10_list.csv")
    cap41_map  = _load_csv_part_map(ref / "capacitors_41_list.csv")
    ind2s_map  = _load_csv_part_map(ref / "inductors_2s_list.csv")

    # Merged capacitor map – 10-type entries take precedence for size
    cap_merged: dict[str, dict] = {}
    for pn, sz in cap41_map.items():
        cap_merged[pn] = {"size": sz, "source": "41-type"}
    for pn, sz in cap10_map.items():
        cap_merged[pn] = {"size": sz, "source": "10-type"}

    ind_merged: dict[str, dict] = {
        pn: {"size": sz} for pn, sz in ind2s_map.items()
    }

    cap10_only: dict[str, dict] = {
        pn: {"size": sz} for pn, sz in cap10_map.items()
    }

    sections = [
        (f"Capacitors ({len(cap_merged)} types)", cap_merged, True),
        ("Inductors (2S)", ind_merged, False),
        (f"Capacitors 10-type ({len(cap10_only)} types)", cap10_only, False),
        (f"Inductors 2S ({len(ind_merged)} types)", ind_merged, False),
    ]

    # -- Sheet title --
    NUM_COLS = 6
    ws.merge_cells(f"A1:{get_column_letter(NUM_COLS)}1")
    title = ws["A1"]
    title.value = "Managed Parts — Usage Frequency"
    title.font = Font(bold=True, size=13)
    title.alignment = Alignment(horizontal="center")

    current_row = 3
    _SUBHDR_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    _USED_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    _ZERO_FILL   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _USED_FONT   = Font(color="006100")
    _ZERO_FONT   = Font(color="A0A0A0", italic=True)

    for section_title, part_meta, has_source in sections:
        if not part_meta:
            continue

        usage = _count_usage(set(part_meta.keys()), components_top, components_bot)

        total_on_board = sum(v["total"] for v in usage.values())
        used_types     = sum(1 for v in usage.values() if v["total"] > 0)

        # -- Section header (blue bar) --
        hdr = ws.cell(row=current_row, column=1, value=section_title)
        hdr.font = Font(bold=True, size=11, color="FFFFFF")
        for col in range(1, NUM_COLS + 1):
            ws.cell(row=current_row, column=col).fill = _HEADER_FILL
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=NUM_COLS
        )
        hdr.alignment = Alignment(horizontal="left")
        current_row += 1

        # -- Summary row --
        summary_text = (
            f"Managed types: {len(part_meta)}    "
            f"Types found on board: {used_types}    "
            f"Total placements: {total_on_board}"
        )
        summary = ws.cell(row=current_row, column=1, value=summary_text)
        summary.font = Font(bold=True)
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=NUM_COLS
        )
        current_row += 1

        # -- Column headers --
        col_headers = ["part_name", "size", "TOP", "BOTTOM", "Total"]
        if has_source:
            col_headers.append("source")
        for col, h in enumerate(col_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = _SUBHDR_FILL
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # -- Data rows: sort by total descending, then part_name --
        sorted_parts = sorted(
            part_meta.items(),
            key=lambda kv: (-usage[kv[0]]["total"], kv[0])
        )

        for pn, meta in sorted_parts:
            cnt = usage[pn]
            is_used = cnt["total"] > 0
            row_fill = _USED_FILL if is_used else _ZERO_FILL
            row_font = _USED_FONT if is_used else _ZERO_FONT

            values = [pn, meta.get("size", ""), cnt["top"], cnt["bottom"], cnt["total"]]
            if has_source:
                values.append(meta.get("source", ""))

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.fill = row_fill
                cell.font = row_font
                cell.border = _THIN_BORDER
                if col in (3, 4, 5):
                    cell.alignment = Alignment(horizontal="center")
            current_row += 1

        current_row += 1  # blank row between sections

    _auto_fit_columns(ws)


def _auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    max_line = max(len(line) for line in lines)
                    max_length = max(max_length, max_line)
        ws.column_dimensions[col_letter].width = min(50, max(10, max_length + 2))
