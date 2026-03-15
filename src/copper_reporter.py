"""Excel report generator for copper ratio analysis."""

from __future__ import annotations

from pathlib import Path
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

from src.models import MatrixLayer


# Style constants
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Conditional fill colors for sub-section grid
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_GREY_FILL = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")


def generate_copper_report(
    layer_results: list[dict],
    copper_data: dict[str, float],
    all_matrix_layers: list[MatrixLayer],
    output_path: str | Path,
) -> None:
    """Generate an Excel report with copper ratio analysis.

    Args:
        layer_results: List of dicts (one per signal layer) with keys:
            - layer_name: str
            - total_ratio: float (0–1)
            - subsection_ratios: np.ndarray (5, 5) or None
            - thickness_mm: float or None
            - image_path: Path or None
        copper_data: dict[layer_name, thickness_mm] for all layers (signal + dielectric)
        all_matrix_layers: list of MatrixLayer objects sorted by row
        output_path: Path to write the .xlsx file
    """
    output_path = Path(output_path)
    excel_dir = output_path.parent
    wb = Workbook()

    # Create Summary sheet
    _create_summary_sheet(wb, layer_results, copper_data, all_matrix_layers)

    # Create per-layer sheets
    for result in layer_results:
        _create_layer_sheet(wb, result, excel_dir=excel_dir)

    # Save
    excel_dir.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Copper report saved: {output_path}")


def _create_summary_sheet(
    wb: Workbook,
    layer_results: list[dict],
    copper_data: dict[str, float],
    all_matrix_layers: list[MatrixLayer],
) -> None:
    """Create the Summary sheet with copper ratios and thickness info."""
    ws = wb.active
    ws.title = "Summary"

    current_row = 1

    # ---- Table A: Copper Ratios by Layer ----
    ws.cell(row=current_row, column=1, value="Copper Ratio Analysis").font = Font(bold=True, size=12)
    current_row += 1

    # Header row for Table A
    header_row = current_row
    ws.cell(row=header_row, column=1, value="Layer Name")
    ws.cell(row=header_row, column=2, value="Total Copper (%)")

    # Style header row
    for col_idx in range(1, 3):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    current_row += 1

    # Data rows for Table A
    for result in layer_results:
        layer_name = result["layer_name"]
        total_ratio = result["total_ratio"]

        # Layer name
        ws.cell(row=current_row, column=1, value=layer_name).border = _THIN_BORDER

        # Total copper ratio
        total_cell = ws.cell(row=current_row, column=2)
        if total_ratio is not None:
            total_cell.value = total_ratio
            total_cell.number_format = "0.00%"
        total_cell.border = _THIN_BORDER

        current_row += 1

    current_row += 2  # Leave a blank row before Table B

    # ---- Table B: Layer Thickness ----
    ws.cell(row=current_row, column=1, value="Layer Thickness Information").font = Font(bold=True, size=12)
    current_row += 1

    # Header row for Table B
    header_row = current_row
    ws.cell(row=header_row, column=1, value="Layer Name")
    ws.cell(row=header_row, column=2, value="Type")
    ws.cell(row=header_row, column=3, value="Thickness (mm)")

    # Style header row
    for col_idx in range(1, 4):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    current_row += 1

    # Data rows for Table B (all layers from matrix_layers that have copper_data)
    total_thickness = 0.0
    for ml in all_matrix_layers:
        if ml.name not in copper_data:
            continue

        thickness = copper_data[ml.name]
        total_thickness += thickness

        ws.cell(row=current_row, column=1, value=ml.name).border = _THIN_BORDER
        ws.cell(row=current_row, column=2, value=ml.type).border = _THIN_BORDER

        thickness_cell = ws.cell(row=current_row, column=3)
        thickness_cell.value = thickness
        thickness_cell.number_format = "0.0000"
        thickness_cell.border = _THIN_BORDER

        current_row += 1

    # Total row
    ws.cell(row=current_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=current_row, column=1).border = _THIN_BORDER
    ws.cell(row=current_row, column=2).border = _THIN_BORDER

    total_cell = ws.cell(row=current_row, column=3)
    total_cell.value = total_thickness
    total_cell.number_format = "0.0000"
    total_cell.font = Font(bold=True)
    total_cell.border = _THIN_BORDER

    # Auto-fit columns
    _auto_fit_columns(ws)


def _create_layer_sheet(wb: Workbook, result: dict, excel_dir: Path = None) -> None:
    """Create a per-layer sheet with details and sub-section grid.

    Args:
        wb: Workbook
        result: Layer result dict
        excel_dir: Directory where Excel file will be saved (for resolving image paths)
    """
    layer_name = result["layer_name"]

    # Sanitize sheet name (Excel has restrictions on sheet names)
    safe_name = (
        layer_name
        .replace("/", "_")
        .replace("\\", "_")
        .replace(":", "_")
        .replace("[", "_")
        .replace("]", "_")
        .replace("*", "_")
        .replace("?", "_")
    )
    # Limit to 31 characters
    safe_name = safe_name[:31]

    ws = wb.create_sheet(safe_name)

    current_row = 1

    # Header info
    ws.cell(row=current_row, column=1, value="Layer:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=layer_name)
    current_row += 1

    ws.cell(row=current_row, column=1, value="Copper Ratio:").font = Font(bold=True)
    ratio_cell = ws.cell(row=current_row, column=2)
    if result["total_ratio"] is not None:
        ratio_cell.value = result["total_ratio"]
        ratio_cell.number_format = "0.00%"
    current_row += 1

    ws.cell(row=current_row, column=1, value="Thickness (mm):").font = Font(bold=True)
    thickness_cell = ws.cell(row=current_row, column=2)
    if result.get("thickness_mm") is not None:
        thickness_cell.value = result["thickness_mm"]
        thickness_cell.number_format = "0.0000"
    else:
        thickness_cell.value = "N/A"
    current_row += 2

    # Sub-section grid
    if result["subsection_ratios"] is not None:
        ws.cell(row=current_row, column=1, value="Sub-section Grid (5×5):").font = Font(bold=True)
        current_row += 1

        # Column headers
        ws.cell(row=current_row, column=1, value="")
        for j in range(1, 6):
            cell = ws.cell(row=current_row, column=j + 1)
            cell.value = f"C{j}"
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER
        current_row += 1

        # Data rows
        ratios = result["subsection_ratios"]
        for i in range(5):
            # Row label
            row_label_cell = ws.cell(row=current_row, column=1)
            row_label_cell.value = f"R{i + 1}"
            row_label_cell.fill = _HEADER_FILL
            row_label_cell.font = _HEADER_FONT
            row_label_cell.border = _THIN_BORDER

            # Data cells
            for j in range(5):
                cell = ws.cell(row=current_row, column=j + 2)
                ratio = ratios[i, j]

                if not np.isnan(ratio):
                    cell.value = ratio
                    cell.number_format = "0.00%"

                    # Conditional fill
                    if ratio > 0.5:
                        cell.fill = _GREEN_FILL
                    elif ratio >= 0.3:
                        cell.fill = _YELLOW_FILL
                    else:
                        cell.fill = _RED_FILL
                else:
                    cell.value = "N/A"
                    cell.fill = _GREY_FILL

                cell.alignment = Alignment(horizontal="center")
                cell.border = _THIN_BORDER

            current_row += 1

        current_row += 1

    # Embedded image
    if result.get("image_path") and excel_dir:
        current_row += 1
        img_path = result["image_path"]

        # Convert relative path to absolute if needed
        if not Path(img_path).is_absolute():
            img_path = excel_dir / img_path

        # Only embed if file exists
        if Path(img_path).exists():
            try:
                # Add label
                ws.cell(row=current_row, column=1, value="Visualization:").font = Font(bold=True)
                current_row += 1

                # Insert image (scale to reasonable size: ~4 inches wide)
                img = XlImage(str(img_path))
                img.width = 288  # pixels at 72 DPI = 4 inches
                img.height = int(288 * img.height / img.width) if img.width else 216  # maintain aspect ratio
                ws.add_image(img, f"A{current_row}")
                current_row += int(img.height / 15) + 2  # Approximate rows needed for image
            except Exception:
                # If image embedding fails, just skip it
                pass

    # Auto-fit columns
    _auto_fit_columns(ws)


def _auto_fit_columns(ws) -> None:
    """Auto-fit column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    max_line = max(len(line) for line in lines)
                    max_length = max(max_length, max_line)
        ws.column_dimensions[col_letter].width = min(50, max(10, max_length + 2))
