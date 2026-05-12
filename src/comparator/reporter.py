"""Excel report generator for ODB++ revision comparison results."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.comparator.base import ComparisonResult, SheetConfig


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

_SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_SECTION_FONT = Font(bold=True, size=11)

# Component change styles
_ADDED_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
_ADDED_FONT = Font(color="006080", bold=True)

_REMOVED_FILL = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
_REMOVED_FONT = Font(color="8B0000", bold=True)

_RELOCATED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_RELOCATED_FONT = Font(color="806000", bold=True)

_MODIFIED_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_MODIFIED_FONT = Font(color="375623", bold=True)

# Checklist transition styles
_FIXED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FIXED_FONT = Font(color="006100", bold=True)

_REGRESSED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_REGRESSED_FONT = Font(color="9C0006", bold=True)

_STILL_FAIL_FILL = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
_STILL_FAIL_FONT = Font(color="8B0000")

_STILL_PASS_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_STILL_PASS_FONT = Font(color="808080")

_NEW_RULE_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
_NEW_RULE_FONT = Font(color="006080")

_REMOVED_RULE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_REMOVED_RULE_FONT = Font(color="808080", italic=True)

# Individual status cell styles (for Old Status / New Status columns)
_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_PASS_FONT = Font(color="006100", bold=True)
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FAIL_FONT = Font(color="9C0006", bold=True)

# Maps for row-level color coding
_CHANGE_STYLE = {
    "ADDED":    (_ADDED_FILL, _ADDED_FONT),
    "REMOVED":  (_REMOVED_FILL, _REMOVED_FONT),
    "RELOCATED": (_RELOCATED_FILL, _RELOCATED_FONT),
    "MODIFIED": (_MODIFIED_FILL, _MODIFIED_FONT),
}

_TRANSITION_STYLE = {
    "FIXED":        (_FIXED_FILL, _FIXED_FONT),
    "REGRESSED":    (_REGRESSED_FILL, _REGRESSED_FONT),
    "STILL_FAIL":   (_STILL_FAIL_FILL, _STILL_FAIL_FONT),
    "STILL_PASS":   (_STILL_PASS_FILL, _STILL_PASS_FONT),
    "NEW_RULE":     (_NEW_RULE_FILL, _NEW_RULE_FONT),
    "REMOVED_RULE": (_REMOVED_RULE_FILL, _REMOVED_RULE_FONT),
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_comparison_report(
    results: list[ComparisonResult],
    output_path: str | Path,
    old_job_name: str = "",
    new_job_name: str = "",
) -> None:
    """Generate Excel comparison report from ComparisonResult list.

    Creates one workbook with an Overview sheet followed by all data sheets
    from each comparator result.
    """
    wb = Workbook()
    _create_overview_sheet(wb, results, old_job_name, new_job_name)

    for comp_result in results:
        for sheet_cfg in comp_result.sheet_configs:
            _create_data_sheet(wb, sheet_cfg)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Comparison report saved: {output_path}")


# ---------------------------------------------------------------------------
# Overview sheet
# ---------------------------------------------------------------------------

def _create_overview_sheet(wb: Workbook, results: list[ComparisonResult],
                           old_job_name: str, new_job_name: str):
    """Create the summary Overview sheet."""
    ws = wb.active
    ws.title = "Overview"

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "ODB++ Revision Comparison Report"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Revision names
    ws["A3"] = "Old Revision:"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = old_job_name

    ws["A4"] = "New Revision:"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = new_job_name

    current_row = 6

    for comp_result in results:
        # Section header
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=4,
        )
        hdr_cell = ws.cell(row=current_row, column=1,
                           value=f"--- {comp_result.title} ---")
        hdr_cell.font = _SECTION_FONT
        hdr_cell.fill = _SECTION_FILL
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).fill = _SECTION_FILL
        current_row += 1

        # Stats from each sheet_config
        for sheet_cfg in comp_result.sheet_configs:
            stats = sheet_cfg.stats
            if not stats:
                continue

            # Component-type stats
            if "layer" in stats:
                layer = stats["layer"]
                added = stats.get("ADDED", 0)
                removed = stats.get("REMOVED", 0)
                relocated = stats.get("RELOCATED", 0)
                modified = stats.get("MODIFIED", 0)
                old_total = stats.get("old_total", 0)
                new_total = stats.get("new_total", 0)

                ws.cell(row=current_row, column=1,
                        value=f"{layer} Layer:").font = Font(bold=True)
                ws.cell(row=current_row, column=2,
                        value=f"Old: {old_total}  /  New: {new_total}")
                current_row += 1

                for label, count, fill, font in [
                    ("Added:", added, _ADDED_FILL, _ADDED_FONT),
                    ("Removed:", removed, _REMOVED_FILL, _REMOVED_FONT),
                    ("Relocated:", relocated, _RELOCATED_FILL, _RELOCATED_FONT),
                    ("Modified:", modified, _MODIFIED_FILL, _MODIFIED_FONT),
                ]:
                    ws.cell(row=current_row, column=2, value=label)
                    cnt_cell = ws.cell(row=current_row, column=3, value=count)
                    cnt_cell.fill = fill
                    cnt_cell.font = font
                    cnt_cell.alignment = Alignment(horizontal="center")
                    current_row += 1
                current_row += 1

            # Checklist-type stats
            elif any(k in stats for k in ("FIXED", "REGRESSED", "STILL_FAIL")):
                for label, key, fill, font in [
                    ("Fixed (FAIL \u2192 PASS):", "FIXED",
                     _FIXED_FILL, _FIXED_FONT),
                    ("Regressed (PASS \u2192 FAIL):", "REGRESSED",
                     _REGRESSED_FILL, _REGRESSED_FONT),
                    ("Still Failing:", "STILL_FAIL",
                     _STILL_FAIL_FILL, _STILL_FAIL_FONT),
                    ("Still Passing:", "STILL_PASS",
                     _STILL_PASS_FILL, _STILL_PASS_FONT),
                ]:
                    ws.cell(row=current_row, column=2, value=label)
                    cnt_cell = ws.cell(row=current_row, column=3,
                                       value=stats.get(key, 0))
                    cnt_cell.fill = fill
                    cnt_cell.font = font
                    cnt_cell.alignment = Alignment(horizontal="center")
                    current_row += 1
                current_row += 1

        current_row += 1

    _auto_fit_columns(ws)


# ---------------------------------------------------------------------------
# Data sheets (generic, driven by SheetConfig)
# ---------------------------------------------------------------------------

def _create_data_sheet(wb: Workbook, cfg: SheetConfig):
    """Create a data sheet from a SheetConfig definition."""
    ws = wb.create_sheet(cfg.sheet_name)

    # Title row
    if cfg.title:
        num_cols = max(len(cfg.columns), 1)
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=num_cols,
        )
        title_cell = ws["A1"]
        title_cell.value = cfg.title
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = Alignment(horizontal="center")

    if not cfg.columns or not cfg.rows:
        if not cfg.rows:
            ws["A3"] = "No changes detected."
        _auto_fit_columns(ws)
        return

    # Detect which column drives row coloring
    color_col = None
    if "Change" in cfg.columns:
        color_col = "Change"
    elif "Transition" in cfg.columns:
        color_col = "Transition"

    # Column headers (row 3)
    header_row = 3
    for col_idx, col_name in enumerate(cfg.columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_offset, row_data in enumerate(cfg.rows, 1):
        current_row = header_row + row_offset

        # Determine row style from the color column
        row_fill = None
        row_font = None
        if color_col:
            val = str(row_data.get(color_col, ""))
            if color_col == "Change":
                style = _CHANGE_STYLE.get(val)
            else:
                style = _TRANSITION_STYLE.get(val)
            if style:
                row_fill, row_font = style

        for col_idx, col_name in enumerate(cfg.columns, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = _THIN_BORDER

            # Apply row-level fill
            if row_fill:
                cell.fill = row_fill

            # Special styling for specific columns
            if col_name == color_col and row_font:
                cell.font = row_font
                cell.alignment = Alignment(horizontal="center")
            elif col_name in ("Old Status", "New Status"):
                val_str = str(value).upper()
                if val_str == "PASS":
                    cell.fill = _PASS_FILL
                    cell.font = _PASS_FONT
                elif val_str == "FAIL":
                    cell.fill = _FAIL_FILL
                    cell.font = _FAIL_FONT
                cell.alignment = Alignment(horizontal="center")

    _auto_fit_columns(ws)


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    max_line = max(len(line) for line in lines)
                    max_length = max(max_length, max_line)
        ws.column_dimensions[col_letter].width = min(50, max(10, max_length + 2))
