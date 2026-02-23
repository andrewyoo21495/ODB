"""
Excel Reporter
Outputs checklist results to a formatted Excel (.xlsx) file using openpyxl.
"""

from typing import List, TYPE_CHECKING
from datetime import datetime

if TYPE_CHECKING:
    pass

from .rule_base import CheckResult, CheckStatus


class ExcelReporter:
    """Generates a formatted Excel report from checklist results."""

    STATUS_FILL_COLORS = {
        'PASS':    '00CC44',
        'FAIL':    'CC0000',
        'WARNING': 'FFAA00',
        'SKIP':    '888888',
    }

    HEADER_FILL = '1F4E79'

    def export(
        self,
        results: List[CheckResult],
        output_path: str,
        pcb_name: str = '',
    ) -> None:
        """
        Write checklist results to an Excel file.

        Args:
            results: List of CheckResult objects.
            output_path: Path to write the .xlsx file.
            pcb_name: Board/product name used in the summary sheet.
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError as e:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            ) from e

        wb = openpyxl.Workbook()

        # ---- Results sheet ----
        ws = wb.active
        ws.title = 'Checklist Results'
        ws.sheet_view.showGridLines = True

        # Column widths
        ws.column_dimensions['A'].width = 12   # ID
        ws.column_dimensions['B'].width = 45   # Rule Name
        ws.column_dimensions['C'].width = 12   # Result
        ws.column_dimensions['D'].width = 55   # Message
        ws.column_dimensions['E'].width = 80   # Details

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor=self.HEADER_FILL)
        header_align = Alignment(horizontal='center', vertical='center')

        headers = ['Rule ID', 'Rule Name', 'Result', 'Message', 'Details']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.row_dimensions[1].height = 22

        for row_idx, result in enumerate(results, 2):
            ws.cell(row_idx, 1, result.rule_id)
            ws.cell(row_idx, 2, result.rule_name)

            status_val = result.status.value
            status_cell = ws.cell(row_idx, 3, status_val)
            fill_color = self.STATUS_FILL_COLORS.get(status_val, '888888')
            status_cell.fill = PatternFill('solid', fgColor=fill_color)
            status_cell.font = Font(bold=True, color='FFFFFF')
            status_cell.alignment = Alignment(horizontal='center')

            ws.cell(row_idx, 4, result.message)

            if result.details:
                details_text = '\n'.join(str(d) for d in result.details[:50])
                if len(result.details) > 50:
                    details_text += f'\n... (+{len(result.details) - 50} more)'
                cell_d = ws.cell(row_idx, 5, details_text)
                cell_d.alignment = Alignment(wrap_text=True)
            else:
                ws.cell(row_idx, 5, '')

        # ---- Summary sheet ----
        ws2 = wb.create_sheet('Summary')
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 15

        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        summary_rows = [
            ('PCB / Product', pcb_name),
            ('Generated', now_str),
            ('Total Rules', str(len(results))),
        ]

        from .rule_base import CheckStatus as CS
        for status in CS:
            count = sum(1 for r in results if r.status == status)
            summary_rows.append((status.value, str(count)))

        bold_font = Font(bold=True, size=11)
        for r_idx, (label, value) in enumerate(summary_rows, 1):
            label_cell = ws2.cell(r_idx, 1, label)
            value_cell = ws2.cell(r_idx, 2, value)
            label_cell.font = bold_font
            # Color-code status rows
            for status in CheckStatus:
                if label == status.value:
                    color = self.STATUS_FILL_COLORS.get(status.value, 'FFFFFF')
                    value_cell.fill = PatternFill('solid', fgColor=color)
                    value_cell.font = Font(bold=True, color='FFFFFF')

        wb.save(output_path)
        print(f"Report saved: {output_path}")
